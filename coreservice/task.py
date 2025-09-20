import secrets
from io import BytesIO

import imgkit
import requests
from django.contrib.humanize.templatetags.humanize import intcomma
from django.core.files import File
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage, FileSystemStorage
from django.core.signing import TimestampSigner
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django_countries import countries
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from weasyprint import HTML
from coreservice.mail_helper import MailSender
from coreservice.models import Booking, Exhibitor, OTP, Invoice
import sefservices.settings as env_variable
import os


# Sending Signal for new registered Exhibitor
def exhibitor_account_activation_mail(instance):
    exhibitor = Exhibitor.objects.get(pk=instance.id)

    print("red")

    code = secrets.token_hex(40)
    signer = TimestampSigner()

    OTP.objects.create(
        email=exhibitor.email,
        code=str(code)
    )

    token = signer.sign(str(code))

    password_setting_link = f'{env_variable.SITE_URL}{reverse("exhibitor:mot_de_pass_creation")}?token={token}&email={exhibitor.email}'
    mail_sender = MailSender()
    mail_sender.exhibitor_password_creation_mail(**{"exhibitor_id": exhibitor.id, "link": password_setting_link})


def exhibitor_password_reset_mail(instance):
    exhibitor = Exhibitor.objects.get(pk=instance.id)

    code = secrets.token_hex(40)
    signer = TimestampSigner()

    OTP.objects.create(
        email=exhibitor.email,
        code=str(code)
    )

    token = str(signer.sign(str(code)))

    password_setting_link = f'{env_variable.SITE_URL}{reverse("exhibitor:reinitialisation_mot_de_passe")}?token={token}&email={exhibitor.email}'
    mail_sender = MailSender()
    mail_sender.exhibitor_password_reset_mail(**{"exhibitor_id": exhibitor.id, "link": password_setting_link})


def exhibitor_booking_submission_mail(instance):
    mail_sender = MailSender()
    mail_sender.exhibitor_booking_submission_mail(**{"booking": instance.id})


# Function Sending Booking Approve mail
def booking_approval_receiver(instance):
    booking = Booking.objects.get(pk=instance.id)
    booking_detail_link = f"{env_variable.SITE_URL}{reverse('exhibitor:reservation_details_exhibitor', kwargs={'booking_ref': booking.booking_ref})}"

    mail_sender = MailSender()
    mail_sender.booking_approval_mail(**{"booking_id": booking.id, 'link': booking_detail_link})


# Function Sending Booking rejection mail
def booking_rejection_receiver(instance):
    booking = Booking.objects.get(pk=instance.id)
    booking_detail_link = f"{env_variable.SITE_URL}{reverse('exhibitor:reservation_details_exhibitor', kwargs={'booking_ref': booking.booking_ref})}"
    mail_sender = MailSender()
    mail_sender.booking_rejection_mail(**{"booking_id": booking.id, 'link': booking_detail_link})


# Payment Received Mail Function
def booking_payment_received_receiver(instance):
    mail_sender = MailSender()
    mail_sender.booking_payment_received_mai(**{"payment": instance.id})


# Admin Payment Notification
def notify_admins_with_permission(**kwargs):
    mail_sender = MailSender()
    mail_sender.transaction_revision_mail(**{"payment": kwargs.get('payment'), 'staff': kwargs.get('staff_id')})


def export_invoices(**kwargs):
    invoices = kwargs.get('invoices')
    staff = kwargs.get('staff')

    # Creation du fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Liste des Factures"

    # Entête de la feuille excel
    headers = ['Référence', 'Référence Réservation', 'Exposant', 'Montant total facturé (FCFA)', 'Total Payé (FCFA)',
               'Montant dû (FCFA)', 'Date émise', 'Status']

    ws.append(headers)

    # Écriture des données sur le fichier excel
    for invoice in invoices:
        if invoice.status == "Paid":
            status = "Payé"
        elif invoice.status == "Partially Paid":
            status = "Paiement Partiel"
        else:
            status = "Non Payé"
        ws.append([
            invoice.invoice_number,
            invoice.booking.booking_ref,
            str(invoice.booking.exhibitor.designation).upper(),
            intcomma(int(invoice.amount_total)),
            intcomma(int(invoice.amount_paid)),
            intcomma(int(invoice.remaining_balance)),
            invoice.issued_at.strftime("%d-%m-%Y"),
            status
        ])

    # Ajustement des colones
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    # Enregistrement temporaire sur le système
    file_name = f"Factures_{timezone.now().strftime('%d%m%Y_%H%M%S')}.xlsx"
    local_dir = 'medias/importations/'

    # Créer le répertoire s'il n'existe pas
    os.makedirs(local_dir, exist_ok=True)

    # Sauvegarde du fichier excel
    local_fs = FileSystemStorage(location=local_dir)
    full_path = local_fs.path(file_name)
    wb.save(full_path)

    # Téléversement du fichier vers le serveur de stockage
    with open(full_path, 'rb') as f:
        django_file = File(f)
        minio_path = f'Exportations/Factures/{file_name}'
        default_storage.save(minio_path, django_file)

    # Suppression du fichier local
    if os.path.exists(full_path):
        os.remove(full_path)

    # Récupération de l'URL du fichier téléversé
    if default_storage.exists(minio_path):
        try:
            url = default_storage.url(minio_path)
        except NotImplementedError:
            url = f"/media/{minio_path}"  # fallback if using file backend

    else:
        url = None

    # Envoi de mail une fois l'exportation terminée
    mail_sender = MailSender()
    mail_sender.invoices_exportation_mail(**{'staff': staff, 'file_url': f"{url}"})
    # if default_storage.exists(minio_path):
    #    default_storage.delete(minio_path)

    print(f"[EXPORT] Invoice export completed. File '{file_name}' sent via email and deleted from Storage Server.")


def export_exhibitor(**kwargs):
    exhibitors = kwargs.get('exhibitors')
    staff = kwargs.get('staff')

    # Creation du fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Liste des exposants"

    # Entête de la feuille excel
    headers = ['Date', 'Nom', 'Prénom', 'Nom de la société', 'Email', 'Téléphone', 'Pays']

    ws.append(headers)

    # Écriture des données sur le fichier excel
    for exhibitor in exhibitors:
        ws.append([
            exhibitor.created_at.strftime("%d-%m-%Y"),
            str(exhibitor.lastname).upper(),
            str(exhibitor.firstname).title(),
            str(exhibitor.designation).upper(),
            exhibitor.email,
            exhibitor.phone,
            str(dict(countries)[exhibitor.country]),
        ])

    # Ajustement des colones
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    # Enregistrement temporaire sur le système
    file_name = f"Liste_Exposant_{timezone.now().strftime('%d%m%Y_%H%M%S')}.xlsx"
    local_dir = 'medias/importations/'

    # Créer le répertoire s'il n'existe pas
    os.makedirs(local_dir, exist_ok=True)

    # Sauvegarde du fichier excel
    local_fs = FileSystemStorage(location=local_dir)
    full_path = local_fs.path(file_name)
    wb.save(full_path)

    # Téléversement du fichier vers le serveur de stockage
    with open(full_path, 'rb') as f:
        django_file = File(f)
        minio_path = f'Exportations/Exposants/{file_name}'
        default_storage.save(minio_path, django_file)

    # Suppression du fichier local
    if os.path.exists(full_path):
        os.remove(full_path)

    # Récupération de l'URL du fichier téléversé
    if default_storage.exists(minio_path):
        try:
            url = default_storage.url(minio_path)
        except NotImplementedError:
            url = f"/media/{minio_path}"  # fallback if using file backend

    else:
        url = None

    # Envoi de mail une fois l'exportation terminée
    mail_sender = MailSender()
    mail_sender.invoices_exportation_mail(**{'staff': staff, 'file_url': f"{url}"})
    # if default_storage.exists(minio_path):
    #    default_storage.delete(minio_path)

    print(f"[EXPORT] Invoice export completed. File '{file_name}' sent via email and deleted from Storage Server.")


def quitus_availability_mail(**kwargs):

    invoice = Invoice.objects.get(invoice_number=kwargs.get('invoice'))
    exhibitor = invoice.booking.exhibitor

    mail_sender = MailSender()
    mail_sender.quitus_mail(booking=invoice.booking.id, exhibitor=exhibitor.id)


def generate_ticket_pdf(ticket):
    """ Generates a PDF ticket using WeasyPrint and uploads to MinIO """

    context = {
        "ticket": ticket,
        "ticket_ref": f"Ticket_{ticket.ticket_ref}"
    }

    html_content = render_to_string("tickets/ticket_pdf_template.html", context)

    # Configure wkhtmltoimage
    options = {
        'format': 'jpg',
        'crop-w': '1000',  # Width in pixels
        'crop-h': '1502',  # Height in pixels
        'quality': '50',  # Not relevant for PNG but no harm
        'encoding': "UTF-8",
        'disable-smart-width': '',
        'custom-header': [],
    }
    if env_variable.PRODUCTION:
        config = imgkit.config(wkhtmltoimage='/usr/bin/wkhtmltoimage')
    else:
        wkhtmltoimage_path = 'C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltoimage.exe'  # Windows
        config = imgkit.config(wkhtmltoimage=wkhtmltoimage_path)

    # Convert HTML to image
    image_io = imgkit.from_string(html_content, False, options=options, config=config)

    # Save image to MinIO
    pdf_filename = f"tickets/pdf/{ticket.ticket_ref}.jpg"
    default_storage.save(pdf_filename, ContentFile(image_io))
    ticket.pdf_ticket = pdf_filename
    ticket.save()
